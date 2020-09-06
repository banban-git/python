import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles.borders import Border, Side

# 新規作成
book = openpyxl.Workbook()
# シート設定
sheet = book.worksheets[0]
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 150
#ヘッダ設定
sheet.cell(row=1, column=1).value = "No"
sheet.cell(row=1, column=2).value = "タイトル"
# set border (black thin line)
side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)

#投稿記事すべて
page_Url = "https://2chmm.com/"
r = requests.get(page_Url)
r.encoding = "utf-8"
soup = BeautifulSoup(r.text, features="html.parser")

count = 1
for entry in soup.find_all("li", class_="entry"):
    for entry in entry.find_all("a"):
        text = entry.text
        url = entry.get("href")
        if (len(text) == 0): continue

        count = count + 1
        sheet.cell(row=count, column=1).value = count -1
        sheet.cell(row=count, column=2).value = '=HYPERLINK("' + url + '" , "' + text.replace('”', '\'').replace('\"', '\'') + '")'
        sheet.cell(row=count, column=2).font = openpyxl.styles.fonts.Font(color='1122cc')
        sheet.cell(row=count, column=1).border = border
        sheet.cell(row=count, column=2).border = border
# 保存
book.save('まとめサイトＵＲＬ.xlsx')
# 終了
book.close()